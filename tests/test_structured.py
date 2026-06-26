"""Tests for structured mode (Phase 6)."""
import json
import shutil
from pathlib import Path

import pandas as pd
import pytest

from pseudoswapper.modes.structured import _resolve_anchor, redact_structured
from pseudoswapper.restore import restore_file
from pseudoswapper.session import clear_session, session_exists
from tests.conftest import make_config

FIXTURE_CSV = Path("tests/fixtures/sample_structured.csv")
FIXTURE_JSON = Path("tests/fixtures/sample_structured.json")
FIXTURE_XLSX = Path("tests/fixtures/sample_structured.xlsx")

CORRELATED_CFG = make_config(
    structured={
        "anchor_field": "employee_id",
        "correlated_fields": ["full_name", "email", "username"],
    }
)


@pytest.fixture(autouse=True)
def _clean(tmp_path):
    yield
    clear_session(tmp_path)


# ── Anchor resolution ────────────────────────────────────────────────────────

def test_anchor_from_cli_arg():
    cols = ["employee_id", "full_name", "email"]
    assert _resolve_anchor(cols, "full_name", make_config()) == "full_name"


def test_anchor_cli_beats_config():
    cfg = make_config(structured={"anchor_field": "employee_id", "correlated_fields": []})
    cols = ["employee_id", "full_name", "email"]
    assert _resolve_anchor(cols, "full_name", cfg) == "full_name"


def test_anchor_from_config():
    cfg = make_config(structured={"anchor_field": "employee_id", "correlated_fields": []})
    cols = ["employee_id", "full_name", "email"]
    assert _resolve_anchor(cols, None, cfg) == "employee_id"


def test_anchor_config_beats_auto_detect():
    # "full_name" would be auto-detected, but config says "employee_id"
    cfg = make_config(structured={"anchor_field": "employee_id", "correlated_fields": []})
    cols = ["employee_id", "full_name", "email"]
    assert _resolve_anchor(cols, None, cfg) == "employee_id"


def test_anchor_auto_detect_full_name():
    cols = ["full_name", "email", "department"]
    assert _resolve_anchor(cols, None, make_config()) == "full_name"


def test_anchor_auto_detect_user_id():
    cols = ["user_id", "email", "department"]
    assert _resolve_anchor(cols, None, make_config()) == "user_id"


def test_anchor_returns_none_when_no_match():
    cols = ["department", "role", "level"]
    assert _resolve_anchor(cols, None, make_config()) is None


# ── CSV processing ───────────────────────────────────────────────────────────

def test_csv_output_has_redacted_suffix(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert out.name == "data.redacted.csv"
    assert out.exists()


def test_anchor_field_tokenized(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    df = pd.read_csv(out, dtype=str, keep_default_na=False)
    assert not any(df["employee_id"].str.startswith("E00"))
    assert all(df["employee_id"].str.contains(r"\[PERSON_\d+\]") | (df["employee_id"] == ""))


def test_same_anchor_value_same_token_across_rows(tmp_path):
    # Rows 0 and 2 both have employee_id "E001"
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    df = pd.read_csv(out, dtype=str, keep_default_na=False)
    assert df["employee_id"].iloc[0] == df["employee_id"].iloc[2]


def test_different_anchors_get_distinct_tokens(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    df = pd.read_csv(out, dtype=str, keep_default_na=False)
    assert df["employee_id"].iloc[0] != df["employee_id"].iloc[1]


def test_correlated_email_registered_as_email_person_token(tmp_path):
    src = tmp_path / "data.csv"
    src.write_text("employee_id,email\nE001,john.doe@acme.com\n")
    cfg = make_config(structured={"anchor_field": "employee_id", "correlated_fields": ["email"]})
    out = redact_structured(src, cfg, tmp_path)
    df = pd.read_csv(out, dtype=str, keep_default_na=False)
    assert "[EMAIL_PERSON_" in df["email"].iloc[0]


def test_null_anchor_row_does_not_crash(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert out.exists()


def test_null_anchor_row_correlated_fields_tokenized_independently(tmp_path):
    src = tmp_path / "data.csv"
    src.write_text("employee_id,email\n,orphan@acme.com\n")
    cfg = make_config(structured={"anchor_field": "employee_id", "correlated_fields": ["email"]})
    out = redact_structured(src, cfg, tmp_path)
    df = pd.read_csv(out, dtype=str, keep_default_na=False)
    assert "[EMAIL_" in df["email"].iloc[0]
    assert "orphan@acme.com" not in df["email"].iloc[0]


def test_session_created_after_structured_redact(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert session_exists(tmp_path)


# ── JSON processing ──────────────────────────────────────────────────────────

def test_json_output_has_redacted_suffix(tmp_path):
    src = shutil.copy(FIXTURE_JSON, tmp_path / "data.json")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert out.name == "data.redacted.json"
    assert out.exists()


def test_json_anchor_tokenized(tmp_path):
    src = shutil.copy(FIXTURE_JSON, tmp_path / "data.json")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    rows = json.loads(out.read_text())
    assert not any(r.get("employee_id", "").startswith("E00") for r in rows if r.get("employee_id"))


def test_json_same_anchor_same_token(tmp_path):
    src = shutil.copy(FIXTURE_JSON, tmp_path / "data.json")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    rows = json.loads(out.read_text())
    # Rows 0 and 2 both have employee_id E001
    r0 = rows[0].get("employee_id")
    r2 = rows[2].get("employee_id")
    assert r0 is not None and r0 == r2


# ── XLSX processing ──────────────────────────────────────────────────────────

def test_xlsx_output_has_redacted_suffix(tmp_path):
    src = shutil.copy(FIXTURE_XLSX, tmp_path / "data.xlsx")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert out.name == "data.redacted.xlsx"
    assert out.exists()


def test_xlsx_anchor_tokenized(tmp_path):
    src = shutil.copy(FIXTURE_XLSX, tmp_path / "data.xlsx")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    df = pd.read_excel(out, dtype=str)
    non_empty = df["employee_id"].dropna()
    assert not any(v.startswith("E00") for v in non_empty)


# ── Equivalent output across formats ────────────────────────────────────────

def test_csv_json_xlsx_produce_equivalent_tokens(tmp_path):
    csv_src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    json_src = shutil.copy(FIXTURE_JSON, tmp_path / "data.json")
    xlsx_src = shutil.copy(FIXTURE_XLSX, tmp_path / "data.xlsx")

    csv_cwd = tmp_path / "csv_run"
    csv_cwd.mkdir()
    json_cwd = tmp_path / "json_run"
    json_cwd.mkdir()
    xlsx_cwd = tmp_path / "xlsx_run"
    xlsx_cwd.mkdir()

    csv_out = redact_structured(Path(csv_src), CORRELATED_CFG, csv_cwd)
    json_out = redact_structured(Path(json_src), CORRELATED_CFG, json_cwd)
    xlsx_out = redact_structured(Path(xlsx_src), CORRELATED_CFG, xlsx_cwd)

    df_csv = pd.read_csv(csv_out, dtype=str, keep_default_na=False)
    json_rows = json.loads(json_out.read_text())
    df_json = pd.DataFrame(json_rows, dtype=str).fillna("")
    df_xlsx = pd.read_excel(xlsx_out, dtype=str).fillna("")

    # Token at row 0, employee_id must be same across all three formats
    assert df_csv["employee_id"].iloc[0] == df_json["employee_id"].iloc[0]
    assert df_csv["employee_id"].iloc[0] == df_xlsx["employee_id"].iloc[0]

    # Correlated email token type must be consistent
    assert "[EMAIL_PERSON_" in df_csv["email"].iloc[0]
    assert "[EMAIL_PERSON_" in df_json["email"].iloc[0]


# ── Round-trip ───────────────────────────────────────────────────────────────
# Round-trip tests simulate the realistic usage pattern:
#   redact structured file → AI produces prose output with tokens → restore prose
# We use full_name as the anchor so the canonical reverse value is the human-readable
# name ("John Doe"), not an opaque ID. The employee_id column is not in correlated_fields
# so it is not tokenized and appears verbatim in the output.

_ROUNDTRIP_CFG = make_config(
    structured={
        "anchor_field": "full_name",
        "correlated_fields": ["email"],
    }
)


def test_round_trip_csv(tmp_path):
    src = shutil.copy(FIXTURE_CSV, tmp_path / "data.csv")
    redacted_path = redact_structured(Path(src), _ROUNDTRIP_CFG, tmp_path)

    df = pd.read_csv(redacted_path, dtype=str, keep_default_na=False)
    person_token = df["full_name"].iloc[0]    # e.g. [PERSON_1]
    email_token = df["email"].iloc[0]         # e.g. [EMAIL_PERSON_1]

    # Simulate AI prose output referencing the tokens
    ai_output = tmp_path / "ai_response.txt"
    ai_output.write_text(
        f"Summary: {person_token} is an engineer with address {email_token}."
    )

    restored = restore_file(ai_output, tmp_path)
    content = restored.read_text()

    assert "John Doe" in content
    assert "john.doe@acme.com" in content


def test_round_trip_json(tmp_path):
    src = shutil.copy(FIXTURE_JSON, tmp_path / "data.json")
    redacted_path = redact_structured(Path(src), _ROUNDTRIP_CFG, tmp_path)

    rows = json.loads(redacted_path.read_text())
    person_token = rows[0]["full_name"]
    email_token = rows[0]["email"]

    ai_output = tmp_path / "ai_response.txt"
    ai_output.write_text(
        f"Report: {person_token} can be reached at {email_token}."
    )

    restored = restore_file(ai_output, tmp_path)
    content = restored.read_text()

    assert "John Doe" in content
    assert "john.doe@acme.com" in content


# ── Multi-sheet XLSX ─────────────────────────────────────────────────────────

def _make_multisheet_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    for row in [["employee_id", "full_name", "email"],
                ["E001", "John Doe", "john.doe@acme.com"],
                ["E002", "Jane Smith", "jane@acme.com"]]:
        ws1.append(row)
    ws2 = wb.create_sheet("Events")
    for row in [["employee_id", "action"],
                ["E001", "login"],
                ["E001", "file_access"]]:
        ws2.append(row)
    ws3 = wb.create_sheet("Metadata")
    for row in [["key", "value"],
                ["report_date", "2024-01-15"],
                ["generated_by", "system"]]:
        ws3.append(row)
    wb.save(path)
    return path


def test_multisheet_all_sheets_present_in_output(tmp_path):
    src = _make_multisheet_xlsx(tmp_path / "multi.xlsx")
    out = redact_structured(src, CORRELATED_CFG, tmp_path)
    sheet_names = pd.ExcelFile(out).sheet_names
    assert "Summary" in sheet_names
    assert "Events" in sheet_names
    assert "Metadata" in sheet_names


def test_multisheet_sheet_count_matches_input(tmp_path):
    src = _make_multisheet_xlsx(tmp_path / "multi.xlsx")
    out = redact_structured(src, CORRELATED_CFG, tmp_path)
    assert len(pd.ExcelFile(out).sheet_names) == 3


def test_multisheet_same_entity_same_mask_across_sheets(tmp_path):
    src = _make_multisheet_xlsx(tmp_path / "multi.xlsx")
    out = redact_structured(src, CORRELATED_CFG, tmp_path)
    df_summary = pd.read_excel(out, sheet_name="Summary", dtype=str, keep_default_na=False)
    df_events = pd.read_excel(out, sheet_name="Events", dtype=str, keep_default_na=False)
    # E001 should map to the same token in both sheets
    summary_e001 = df_summary["employee_id"].iloc[0]
    events_e001 = df_events["employee_id"].iloc[0]
    assert summary_e001 == events_e001


def test_multisheet_pii_redacted_in_all_sheets(tmp_path):
    src = _make_multisheet_xlsx(tmp_path / "multi.xlsx")
    out = redact_structured(src, CORRELATED_CFG, tmp_path)
    df_summary = pd.read_excel(out, sheet_name="Summary", dtype=str, keep_default_na=False)
    assert "John Doe" not in df_summary["full_name"].values
    assert "john.doe@acme.com" not in df_summary["email"].values


def test_multisheet_clean_sheet_data_preserved(tmp_path):
    src = _make_multisheet_xlsx(tmp_path / "multi.xlsx")
    out = redact_structured(src, CORRELATED_CFG, tmp_path)
    df_meta = pd.read_excel(out, sheet_name="Metadata", dtype=str, keep_default_na=False)
    assert "report_date" in df_meta["key"].values


def test_single_sheet_xlsx_unaffected(tmp_path):
    src = shutil.copy(FIXTURE_XLSX, tmp_path / "data.xlsx")
    out = redact_structured(Path(src), CORRELATED_CFG, tmp_path)
    assert out.name == "data.redacted.xlsx"
    assert out.exists()
    df = pd.read_excel(out, dtype=str, keep_default_na=False)
    non_empty = df["employee_id"].dropna()
    assert not any(v.startswith("E00") for v in non_empty if v)
